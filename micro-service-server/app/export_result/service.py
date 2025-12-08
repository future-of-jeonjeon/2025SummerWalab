from datetime import datetime
from io import BytesIO
from typing import List, Tuple
from zoneinfo import ZoneInfo

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

import app.contest.repository as contest_repo
import app.submission.repository as submission_repo
from app.contest.models import Contest
from app.export_result import repository as export_repo
from app.user.models import UserData, User


def _format_time(dt: datetime | None) -> str:
    if dt is None:
        return "-"
    try:
        local_dt = dt.astimezone(ZoneInfo("Asia/Seoul"))
    except Exception:
        local_dt = dt
    return local_dt.strftime("%Y-%m-%d %H:%M:%S")


async def build_rank_rows(contest: Contest, db: AsyncSession) -> Tuple[list, dict, dict]:
    problems = await export_repo.fetch_contest_problems(db, contest.id)
    if not problems:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No problems found for contest")

    problem_score_map = export_repo.build_problem_score_map(problems)
    submissions = await export_repo.fetch_contest_submissions(db, contest.id)
    best_scores = export_repo.aggregate_best_scores(submissions, problem_score_map)

    # Score map for total_score from submission repo
    scores_list = await submission_repo.fetch_contest_user_scores(db, contest.id)
    total_score_map = {item["user_id"]: item["total_score"] for item in scores_list}

    # Raw ranks (includes total_time)
    if contest.rule_type == "ACM":
        raw_ranks = await contest_repo.get_acm_contest_rank(contest.id, db)
    else:
        raw_ranks = await contest_repo.get_oi_contest_rank(contest.id, db)

    rows = []
    for rank_row, user, userdata in raw_ranks:
        total_score = total_score_map.get(user.id, 0)
        total_time = getattr(rank_row, "total_time", None)
        # Solve count recomputed from best_scores against full_score
        solved = 0
        user_best = best_scores.get(user.id, {})
        tie_time_candidates = []
        for pid, meta in user_best.items():
            full_score = problem_score_map.get(pid, {}).get("full_score", 0)
            if full_score and meta["score"] >= full_score:
                solved += 1
            # Only consider problems that contributed (>0 score) for tie-breaker
            if meta["score"] > 0 and meta.get("best_time"):
                tie_time_candidates.append(meta["best_time"])

        # Tie-breaker: earliest time when any scored problem was achieved (use latest among scored problems to represent completion)
        tie_time = None
        if tie_time_candidates:
            try:
                tie_time = max(tie_time_candidates)
            except Exception:
                tie_time = None

        rows.append(
            {
                "user_id": user.id,
                "username": user.username,
                "real_name": userdata.name if isinstance(userdata, UserData) else None,
                "student_id": userdata.student_id if isinstance(userdata, UserData) else None,
                "total_score": total_score,
                "total_time": total_time,
                "accepted_number": solved,
                "tie_time": tie_time,
            }
        )

    # Sort: total_score desc, then earliest tie_time (based only on scored problems) asc, fallback to total_time asc, then user_id
    rows.sort(
        key=lambda x: (
            -x["total_score"],
            (x["tie_time"].timestamp() if isinstance(x.get("tie_time"), datetime) else float("inf")),
            float("inf") if x["total_time"] is None else float(x["total_time"]),
            x["user_id"],
        )
    )

    # Return rows, problems, best_scores for further formatting
    return rows, {p.id: p for p in problems}, best_scores


def _add_headers(ws, problems: List, width_map: dict):
    base_headers = ["순위", "이름", "학번", "해결", "점수"]
    # Merge base headers over two rows
    for idx, title in enumerate(base_headers, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        ws.merge_cells(start_row=1, start_column=idx, end_row=2, end_column=idx)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        width_map[idx] = max(width_map.get(idx, 0), len(str(title)))

    col = len(base_headers) + 1
    for p in problems:
        # Merge two columns for problem label
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        title_cell = ws.cell(row=1, column=col, value=str(p._id or p.id))
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True)
        width_map[col] = max(width_map.get(col, 0), len(str(p._id or p.id)))

        ws.cell(row=2, column=col, value="테스트 결과").alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col + 1, value="해당 점수를 받은 최초 시각").alignment = Alignment(horizontal="center")
        width_map[col] = max(width_map.get(col, 0), len("테스트 결과"))
        width_map[col + 1] = max(width_map.get(col + 1, 0), len("해당 점수를 받은 최초 시각"))
        col += 2


def _write_rows(ws, rows, problems_ordered: List, best_scores_map, problem_score_map, width_map: dict):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for idx, entry in enumerate(rows, start=3):
        rank = idx - 2
        solved = entry.get("accepted_number", 0)
        solved_str = f"{solved}/{len(problems_ordered)}"
        base_values = [
            rank,
            entry.get("real_name") or entry.get("username"),
            entry.get("student_id") or "-",
            solved_str,
            entry.get("total_score", 0),
        ]
        col = 1
        for val in base_values:
            cell = ws.cell(row=idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            width_map[col] = max(width_map.get(col, 0), len(str(val)))
            col += 1

        user_best = best_scores_map.get(entry["user_id"], {})
        for p in problems_ordered:
            best = user_best.get(p.id)
            if best:
                test_result = f"{best['passed']}/{best['total']}" if best["total"] else "-"
                time_str = _format_time(best.get("best_time"))
            else:
                test_result = "-"
                time_str = "-"
            for val in (test_result, time_str):
                cell = ws.cell(row=idx, column=col, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
                width_map[col] = max(width_map.get(col, 0), len(str(val)))
                col += 1


async def generate_contest_result_workbook(contest_id: int, db: AsyncSession) -> BytesIO:
    contest = await export_repo.fetch_contest(db, contest_id)
    if not contest:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contest not found")

    rows, problem_map, best_scores = await build_rank_rows(contest, db)
    problems_ordered = list(problem_map.values())
    problem_score_map = export_repo.build_problem_score_map(problems_ordered)

    wb = Workbook()
    ws = wb.active
    ws.title = "contest_result"

    width_map: dict[int, int] = {}
    _add_headers(ws, problems_ordered, width_map)
    _write_rows(ws, rows, problems_ordered, best_scores, problem_score_map, width_map)

    # Adjust column widths a bit
    for col_idx, max_len in width_map.items():
        # Add small padding so text is not cramped
        width = max_len + 2
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
